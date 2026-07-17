from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
import httpx
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from PIL import Image, ImageDraw
import json
import io
import os
import hashlib
import time
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="SIGPAC Sentinel API", version="20.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CACHE_DIR = Path("cache")
CACHE_DIR.mkdir(exist_ok=True)

SIGPAC_CONSULTA_URL  = "https://sigpac-hubcloud.es/servicioconsultassigpac/query"
COPERNICUS_TOKEN_URL = "https://identity.dataspace.copernicus.eu/auth/realms/CDSE/protocol/openid-connect/token"
COPERNICUS_SEARCH_URL = "https://stac.dataspace.copernicus.eu/v1/search"

# Sentinel Hub Process API (requiere token OAuth)
PROCESS_API_URL = "https://sh.dataspace.copernicus.eu/api/v1/process"
# WMS de Copernicus Data Space
WMS_URL = "https://sh.dataspace.copernicus.eu/ogc/wms"

COPERNICUS_USER = os.getenv("COPERNICUS_USER", "")
AEMET_API_KEY  = os.getenv("AEMET_API_KEY", "")
AEMET_BASE     = "https://opendata.aemet.es/opendata/api"
COPERNICUS_PASS = os.getenv("COPERNICUS_PASS", "")
_token_cache = {"token": None, "expires_at": 0}


def cache_key(prefix: str, **kwargs) -> str:
    key = json.dumps(kwargs, sort_keys=True)
    return hashlib.md5(f"{prefix}_{key}".encode()).hexdigest()


async def get_copernicus_token() -> str:
    now = time.time()
    if _token_cache["token"] and now < _token_cache["expires_at"] - 60:
        return _token_cache["token"]
    if not COPERNICUS_USER or not COPERNICUS_PASS:
        raise HTTPException(status_code=500, detail="Credenciales Copernicus no configuradas.")
    async with httpx.AsyncClient() as client:
        resp = await client.post(
            COPERNICUS_TOKEN_URL,
            data={"grant_type": "password", "username": COPERNICUS_USER,
                  "password": COPERNICUS_PASS, "client_id": "cdse-public"},
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        _token_cache["token"] = data["access_token"]
        _token_cache["expires_at"] = now + data.get("expires_in", 3600)
        logger.info("Token Copernicus obtenido")
        return _token_cache["token"]


def geojson_to_mask(geojson: dict, width: int, height: int, bbox: list) -> Image.Image:
    """Crea máscara binaria: blanco=dentro parcela, negro=fuera."""
    min_lon, min_lat, max_lon, max_lat = bbox
    lon_range = max_lon - min_lon
    lat_range = max_lat - min_lat

    def to_px(lon, lat):
        x = int((lon - min_lon) / lon_range * width)
        y = int((max_lat - lat) / lat_range * height)
        return (max(0, min(width-1, x)), max(0, min(height-1, y)))

    mask = Image.new("L", (width, height), 0)
    draw = ImageDraw.Draw(mask)

    for feature in geojson.get("features", []):
        geom = feature.get("geometry", {})
        geom_type = geom.get("type", "")
        if geom_type == "Polygon":
            rings = geom.get("coordinates", [])
            if rings:
                pts = [to_px(c[0], c[1]) for c in rings[0]]
                if len(pts) >= 3:
                    draw.polygon(pts, fill=255)
                for ring in rings[1:]:
                    pts_h = [to_px(c[0], c[1]) for c in ring]
                    if len(pts_h) >= 3:
                        draw.polygon(pts_h, fill=0)
        elif geom_type == "MultiPolygon":
            for polygon in geom.get("coordinates", []):
                if polygon:
                    pts = [to_px(c[0], c[1]) for c in polygon[0]]
                    if len(pts) >= 3:
                        draw.polygon(pts, fill=255)
                    for ring in polygon[1:]:
                        pts_h = [to_px(c[0], c[1]) for c in ring]
                        if len(pts_h) >= 3:
                            draw.polygon(pts_h, fill=0)
    return mask


def aplicar_mascara_jpeg(img_bytes: bytes, geojson: dict, bbox: str) -> bytes:
    """
    Aplica máscara de parcela a imagen JPEG.
    Fuera de parcela → gris oscuro semitransparente para que se distinga.
    Devuelve JPEG.
    """
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    bbox_floats = list(map(float, bbox.split(",")))
    mask = geojson_to_mask(geojson, img.width, img.height, bbox_floats)
    mask_arr = np.array(mask)

    img_arr = np.array(img, dtype=np.uint8)
    # Fuera de parcela: oscurecer al 20% para distinguir sin transparencia
    outside = mask_arr < 128
    img_arr[outside] = (img_arr[outside] * 0.15).astype(np.uint8)

    result = Image.fromarray(img_arr)
    buf = io.BytesIO()
    result.save(buf, format='JPEG', quality=95)
    return buf.getvalue()


def stats_dentro_parcela(img_bytes: bytes, geojson: dict, bbox: str, indice: str) -> dict:
    """Calcula estadísticas SOLO con los píxeles dentro de la parcela."""
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    bbox_floats = list(map(float, bbox.split(",")))
    mask = geojson_to_mask(geojson, img.width, img.height, bbox_floats)
    mask_arr = np.array(mask) > 128

    arr = np.array(img, dtype=np.float32) / 255.0
    verde = arr[:, :, 1]
    pixeles = verde[mask_arr]

    if len(pixeles) == 0:
        return {"indice": indice, "min": 0, "max": 0, "mean": 0, "std": 0}

    return {
        "indice": indice,
        "min": float(pixeles.min()),
        "max": float(pixeles.max()),
        "mean": float(pixeles.mean()),
        "std": float(pixeles.std()),
        "pixeles_parcela": int(mask_arr.sum()),
    }



def bbox_to_float(bbox_str: str):
    return list(map(float, bbox_str.split(",")))


async def procesar_sentinel_evalscript(
    bbox: str,
    fecha: str,
    evalscript: str,
    token: str,
    width: int = 1024,
    height: int = 1024,
) -> Optional[bytes]:
    """
    Usa la Sentinel Hub Process API para obtener imágenes procesadas.
    Esta API SÍ funciona con el token OAuth de Copernicus Data Space.
    """
    min_lon, min_lat, max_lon, max_lat = bbox_to_float(bbox)

    # Fecha inicio y fin (día completo)
    fecha_dt = datetime.strptime(fecha, "%Y-%m-%d")
    fecha_inicio = fecha_dt.strftime("%Y-%m-%dT00:00:00Z")
    fecha_fin = (fecha_dt + timedelta(days=1)).strftime("%Y-%m-%dT00:00:00Z")

    payload = {
        "input": {
            "bounds": {
                "bbox": [min_lon, min_lat, max_lon, max_lat],
                "properties": {"crs": "http://www.opengis.net/def/crs/EPSG/0/4326"}
            },
            "data": [{
                "type": "sentinel-2-l2a",
                "dataFilter": {
                    "timeRange": {"from": fecha_inicio, "to": fecha_fin},
                    "maxCloudCoverage": 80,
                }
            }]
        },
        "output": {
            "width": width,
            "height": height,
            "responses": [{"identifier": "default", "format": {"type": "image/jpeg", "parameters": {"quality": 95}}}]
        },
        "evalscript": evalscript,
    }

    try:
        async with httpx.AsyncClient(
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
                "Accept": "image/jpeg",
            },
            timeout=120,
        ) as client:
            resp = await client.post(PROCESS_API_URL, json=payload)
            logger.info(f"Process API status: {resp.status_code}")
            if resp.status_code == 200:
                return resp.content
            else:
                logger.error(f"Process API error: {resp.text[:300]}")
    except Exception as e:
        logger.error(f"Error Process API: {e}")
    return None


# Evalscripts para cada índice
EVALSCRIPTS = {
    "RGB": """
//VERSION=3
function setup() {
  return { input: ["B04", "B03", "B02"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  return [2.5 * sample.B04, 2.5 * sample.B03, 2.5 * sample.B02];
}
""",
    "NDVI": """
//VERSION=3
function setup() {
  return { input: ["B08", "B04"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let ndvi = (sample.B08 - sample.B04) / (sample.B08 + sample.B04 + 1e-10);
  if (ndvi < -0.5) return [0.05, 0.05, 0.05];
  else if (ndvi < 0) return [0.75, 0.75, 0.75];
  else if (ndvi < 0.1) return [0.86, 0.86, 0.86];
  else if (ndvi < 0.2) return [1, 1, 0.88];
  else if (ndvi < 0.3) return [0.86, 0.96, 0.72];
  else if (ndvi < 0.4) return [0.56, 0.82, 0.54];
  else if (ndvi < 0.5) return [0.27, 0.67, 0.36];
  else if (ndvi < 0.6) return [0.13, 0.52, 0.26];
  else if (ndvi < 0.7) return [0.05, 0.39, 0.16];
  else return [0.0, 0.27, 0.09];
}
""",
    "NDWI": """
//VERSION=3
function setup() {
  return { input: ["B03", "B08"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let ndwi = (sample.B03 - sample.B08) / (sample.B03 + sample.B08 + 1e-10);
  let val = (ndwi + 1) / 2;
  return [1 - val, 1 - val, val];
}
""",
    "EVI": """
//VERSION=3
function setup() {
  return { input: ["B08", "B04", "B02"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let evi = 2.5 * (sample.B08 - sample.B04) / (sample.B08 + 6*sample.B04 - 7.5*sample.B02 + 1 + 1e-10);
  let val = Math.min(Math.max((evi + 1) / 2, 0), 1);
  return [1 - val, val, 1 - val];
}
""",
    "NDRE": """
//VERSION=3
function setup() {
  return { input: ["B08", "B05"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let ndre = (sample.B08 - sample.B05) / (sample.B08 + sample.B05 + 1e-10);
  let val = Math.min(Math.max((ndre + 1) / 2, 0), 1);
  return [1 - val, val, 0.5 - val * 0.5];
}
""",
    "SAVI": """
//VERSION=3
function setup() {
  return { input: ["B08", "B04"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let savi = 1.5 * (sample.B08 - sample.B04) / (sample.B08 + sample.B04 + 0.5 + 1e-10);
  let val = Math.min(Math.max((savi + 1) / 2, 0), 1);
  return [1 - val, val, 0.2];
}
""",
}

INDICES_INFO = {
    "NDVI": "Normalized Difference Vegetation Index",
    "NDWI": "Normalized Difference Water Index",
    "EVI":  "Enhanced Vegetation Index",
    "NDRE": "Normalized Difference Red Edge",
    "SAVI": "Soil-Adjusted Vegetation Index",
}


# ── ENDPOINTS ────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {
        "status": "ok",
        "version": "20.1.0",
        "timestamp": datetime.utcnow().isoformat(),
        "copernicus_configured": bool(COPERNICUS_USER and COPERNICUS_PASS),
    }


@app.get("/sigpac/punto")
async def get_parcela_por_punto(lat: float = Query(...), lon: float = Query(...)):
    ck = cache_key("sigpac_punto", lat=round(lat, 6), lon=round(lon, 6))
    cache_file = CACHE_DIR / f"sigpac_{ck}.geojson"
    if cache_file.exists():
        return JSONResponse(content=json.loads(cache_file.read_text()))

    url = f"{SIGPAC_CONSULTA_URL}/recinfobypoint/4326/{lon}/{lat}.geojson"
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            resp = await client.get(url)
            resp.raise_for_status()
            data = resp.json()
        if not data.get("features"):
            raise HTTPException(status_code=404, detail="No se encontró parcela.")
        cache_file.write_text(json.dumps(data))
        return JSONResponse(content=data)
    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Error SIGPAC: {str(e)}")


@app.get("/sentinel/buscar")
async def buscar_imagenes(
    bbox: str = Query(...),
    fecha_inicio: str = Query(...),
    fecha_fin: str = Query(...),
    max_nubosidad: float = Query(30.0),
):
    try:
        min_lon, min_lat, max_lon, max_lat = map(float, bbox.split(","))
    except ValueError:
        raise HTTPException(status_code=400, detail="bbox invalido")

    # API STAC Copernicus - POST /search (sin token para búsquedas)
    body = {
        "collections": ["sentinel-2-l2a"],
        "bbox": [min_lon, min_lat, max_lon, max_lat],
        "datetime": f"{fecha_inicio}T00:00:00Z/{fecha_fin}T23:59:59Z",
        "limit": 10,
        "sortby": [{"field": "datetime", "direction": "desc"}],
        "query": {"eo:cloud_cover": {"lte": max_nubosidad}},
    }
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                "https://stac.dataspace.copernicus.eu/v1/search",
                json=body
            )
            resp.raise_for_status()
            data = resp.json()
        productos = []
        for item in data.get("features", []):
            props = item.get("properties", {})
            cloud = props.get("eo:cloud_cover")
            fecha_img = props.get("datetime", "")[:10]
            nombre = item.get("id", "")
            productos.append({
                "id": nombre,
                "nombre": nombre,
                "fecha": fecha_img,
                "nubosidad": round(float(cloud), 1) if cloud is not None else None,
                "size_mb": 0.0,
            })
        return {"total": len(productos), "productos": productos}
    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Error Copernicus STAC: {e}")


@app.get("/imagen/rgb")
async def imagen_rgb(
    bbox: str = Query(...),
    fecha: str = Query(..., description="YYYY-MM-DD"),
    geojson: Optional[str] = Query(None, description="GeoJSON parcela para recorte"),
):
    """Imagen color natural recortada por geometría de parcela."""
    ck = cache_key("rgb10", bbox=bbox, fecha=fecha, mask=bool(geojson))
    cache_png = CACHE_DIR / f"{ck}_rgb.jpg"

    if cache_png.exists():
        return StreamingResponse(io.BytesIO(cache_png.read_bytes()), media_type="image/jpeg")

    try:
        token = await get_copernicus_token()
    except HTTPException:
        return _demo_rgb(cache_png)

    img_bytes = await procesar_sentinel_evalscript(bbox, fecha, EVALSCRIPTS["RGB"], token)

    if not img_bytes:
        logger.warning("Process API falló para RGB, usando demo")
        return _demo_rgb(cache_png)

    # Aplicar máscara de parcela si se proporciona GeoJSON
    if geojson:
        try:
            geojson_data = json.loads(geojson)
            img_bytes = aplicar_mascara_jpeg(img_bytes, geojson_data, bbox)
            logger.info("Máscara RGB aplicada correctamente")
        except Exception as e:
            logger.warning(f"Error aplicando máscara RGB: {e}")

    cache_png.write_bytes(img_bytes)
    return StreamingResponse(io.BytesIO(img_bytes), media_type="image/jpeg")


@app.get("/indices/lista")
async def lista_indices():
    return {k: {"descripcion": v, "evalscript": True} for k, v in INDICES_INFO.items()}


@app.get("/indice/calcular")
async def calcular_indice(
    bbox: str = Query(...),
    fecha: str = Query(..., description="YYYY-MM-DD"),
    indice: str = Query(...),
    geojson: Optional[str] = Query(None, description="GeoJSON parcela para recorte"),
    formato: str = Query("png"),
):
    """Calcula índice usando Sentinel Hub Process API con evalscript."""
    indice = indice.upper()
    if indice not in EVALSCRIPTS:
        raise HTTPException(status_code=400, detail=f"Indice desconocido: {list(INDICES_INFO.keys())}")

    ck = cache_key("indice6", bbox=bbox, fecha=fecha, idx=indice)
    cache_png = CACHE_DIR / f"{ck}.jpg"
    cache_stats = CACHE_DIR / f"{ck}_stats.json"

    if cache_png.exists() and formato == "png":
        return StreamingResponse(io.BytesIO(cache_png.read_bytes()), media_type="image/png")
    if cache_stats.exists() and formato == "stats":
        return JSONResponse(content=json.loads(cache_stats.read_text()))

    try:
        token = await get_copernicus_token()
    except HTTPException:
        return _demo_indice_simple(indice, cache_png, cache_stats, formato)

    png_bytes = await procesar_sentinel_evalscript(bbox, fecha, EVALSCRIPTS[indice], token)

    if not png_bytes:
        return _demo_indice_simple(indice, cache_png, cache_stats, formato)

    # Aplicar máscara y calcular estadísticas solo dentro de parcela
    if geojson:
        try:
            geojson_data = json.loads(geojson)
            stats = stats_dentro_parcela(png_bytes, geojson_data, bbox, indice)
            png_bytes = aplicar_mascara_jpeg(png_bytes, geojson_data, bbox)
            logger.info(f"Máscara índice aplicada: {stats.get('pixeles_parcela')} px en parcela")
        except Exception as e:
            logger.warning(f"Error máscara índice: {e}")
            img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
            arr = np.array(img, dtype=np.float32) / 255.0
            verde = arr[:, :, 1]
            stats = {"indice": indice, "min": float(verde.min()), "max": float(verde.max()),
                     "mean": float(verde.mean()), "std": float(verde.std())}
    else:
        img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
        arr = np.array(img, dtype=np.float32) / 255.0
        verde = arr[:, :, 1]
        stats = {"indice": indice, "min": float(verde.min()), "max": float(verde.max()),
                 "mean": float(verde.mean()), "std": float(verde.std())}

    cache_stats.write_text(json.dumps(stats))

    if formato == "stats":
        return JSONResponse(content=stats)

    cache_png.write_bytes(png_bytes)
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/jpeg")


def _demo_rgb(cache_png: Path):
    np.random.seed(123)
    size = (256, 256)
    x, y = np.meshgrid(np.linspace(0, 1, size[1]), np.linspace(0, 1, size[0]))
    base = 0.5 + 0.3 * np.exp(-((x - 0.5)**2 + (y - 0.5)**2) / 0.2)
    r = np.clip(base * 80 + np.random.normal(0, 5, size), 50, 130).astype(np.uint8)
    g = np.clip(base * 120 + np.random.normal(0, 5, size), 80, 180).astype(np.uint8)
    b = np.clip(base * 50 + np.random.normal(0, 5, size), 30, 90).astype(np.uint8)
    rgb = np.stack([r, g, b], axis=2)
    img = Image.fromarray(rgb, mode='RGB')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    png_bytes = buf.read()
    cache_png.write_bytes(png_bytes)
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/jpeg")


def _demo_indice_simple(indice, cache_png, cache_stats, formato):
    np.random.seed(42)
    size = (256, 256)
    x, y = np.meshgrid(np.linspace(0, 1, size[1]), np.linspace(0, 1, size[0]))
    base = 0.3 + 0.4 * np.exp(-((x - 0.5)**2 + (y - 0.5)**2) / 0.15)
    vals = np.clip(base + np.random.normal(0, 0.05, size), 0, 1).astype(np.float32)

    stats = {"indice": indice, "min": float(vals.min()), "max": float(vals.max()),
             "mean": float(vals.mean()), "std": float(vals.std()), "modo": "DEMO"}
    cache_stats.write_text(json.dumps(stats))

    if formato == "stats":
        return JSONResponse(content=stats)

    cmaps = {"NDVI": "RdYlGn", "NDWI": "Blues", "EVI": "YlGn", "NDRE": "RdYlGn", "SAVI": "YlGn"}
    fig, ax = plt.subplots(figsize=(8, 8), dpi=100)
    fig.patch.set_facecolor('#0a0f0d')
    ax.set_facecolor('#0a1a0d')
    im = ax.imshow(vals, cmap=cmaps.get(indice, "RdYlGn"), vmin=0, vmax=1)
    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    ax.set_title(f"{indice} - DEMO", color='#e2ffe8', fontsize=12, fontweight='bold')
    ax.axis('off')
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', dpi=100, facecolor='#0a0f0d')
    plt.close()
    buf.seek(0)
    png_bytes = buf.read()
    cache_png.write_bytes(png_bytes)
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/jpeg")




# Zonas NDVI para mapa de producción
ZONAS_NDVI = [
    {"zona": 1, "label": "Zona 1", "min": 0.90, "max": 1.00, "color": (0,  80,  0)},
    {"zona": 2, "label": "Zona 2", "min": 0.80, "max": 0.89, "color": (0,  120, 0)},
    {"zona": 3, "label": "Zona 3", "min": 0.70, "max": 0.79, "color": (34, 170, 34)},
    {"zona": 4, "label": "Zona 4", "min": 0.60, "max": 0.69, "color": (100,200, 50)},
    {"zona": 5, "label": "Zona 5", "min": 0.50, "max": 0.59, "color": (220,220, 0)},
    {"zona": 6, "label": "Zona 6", "min": 0.40, "max": 0.49, "color": (255,180, 0)},
    {"zona": 7, "label": "Zona 7", "min": 0.30, "max": 0.39, "color": (255,120, 0)},
    {"zona": 8, "label": "Zona 8", "min": 0.20, "max": 0.29, "color": (220, 60, 0)},
    {"zona": 9, "label": "Zona 9", "min": 0.10, "max": 0.19, "color": (200, 30, 30)},
    {"zona":10, "label": "Zona 10","min": 0.00, "max": 0.09, "color": (140,  0,  0)},
]

PIXEL_AREA_M2 = 100       # 10m x 10m = 100 m² por pixel Sentinel-2
M2_PER_HA    = 10000.0   # m² por hectárea


@app.get("/ndvi/zonas")
async def ndvi_zonas(
    bbox: str = Query(...),
    fecha: str = Query(...),
    geojson: Optional[str] = Query(None),
):
    """
    Devuelve imagen NDVI coloreada por zonas de producción
    y estadísticas de superficie por zona.
    """
    ck = cache_key("ndvi_zonas11", bbox=bbox, fecha=fecha, mask=bool(geojson))
    cache_img = CACHE_DIR / f"{ck}_zonas.jpg"
    cache_data = CACHE_DIR / f"{ck}_zonas.json"

    if cache_img.exists() and cache_data.exists():
        return JSONResponse(content={
            "imagen_url": f"/ndvi/zonas/imagen?ck={ck}",
            "zonas": json.loads(cache_data.read_text()),
        })

    try:
        token = await get_copernicus_token()
    except HTTPException:
        raise HTTPException(status_code=500, detail="Error obteniendo token Copernicus")

    # Evalscript que devuelve NDVI en escala de grises (0-255 = NDVI -1 a 1)
    # Evalscript que codifica NDVI en canal R y G como uint8 (0-255)
    # R = parte alta del valor (floor), G = decimales, permite reconstruir NDVI con precisión
    evalscript_ndvi_rgb = """
//VERSION=3
function setup() {
  return { input: ["B08", "B04"], output: { bands: 3 } };
}
function evaluatePixel(sample) {
  let ndvi = (sample.B08 - sample.B04) / (sample.B08 + sample.B04 + 1e-10);
  // Codificar NDVI [-1,1] en dos canales uint8 para mayor precisión
  let v = Math.min(Math.max((ndvi + 1.0) / 2.0, 0), 1);  // [0,1]
  let high = Math.floor(v * 255);
  let low  = Math.floor((v * 255 - high) * 255);
  // Canal B = 128 como marcador para distinguir pixeles validos de fondo
  return [high / 255.0, low / 255.0, 0.5];
}
"""

    raw_bytes = await procesar_sentinel_evalscript(
        bbox, fecha, evalscript_ndvi_rgb, token, width=1024, height=1024
    )

    if not raw_bytes:
        raise HTTPException(status_code=502, detail="No se pudo obtener imagen Sentinel")

    # Decodificar NDVI desde imagen RGB de 2 canales
    img_rgb = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
    rgb_arr = np.array(img_rgb, dtype=np.float32) / 255.0

    # Reconstruir valor [0,1] desde canal R (parte entera) y G (decimales)
    v = rgb_arr[:, :, 0] + rgb_arr[:, :, 1] / 255.0
    # Convertir de [0,1] a NDVI [-1,1]
    ndvi_arr = v * 2.0 - 1.0

    # Marcar píxeles sin datos (canal B ~ 0, no es 0.5)
    valid_mask = rgb_arr[:, :, 2] > 0.25
    ndvi_arr[~valid_mask] = -999  # valor centinela fuera de rango

    # Aplicar máscara de parcela - CRÍTICO: mismo bbox que la imagen
    mask_arr = None
    geojson_data = None
    if geojson:
        try:
            geojson_data = json.loads(geojson)
            bbox_floats = list(map(float, bbox.split(",")))
            # La máscara debe tener exactamente las mismas dimensiones que la imagen NDVI
            mask = geojson_to_mask(geojson_data, img_rgb.width, img_rgb.height, bbox_floats)
            mask_arr = np.array(mask) > 128
            logger.info(f"Máscara creada: {mask_arr.sum()} píxeles dentro de parcela de {img_rgb.width}x{img_rgb.height}")
        except Exception as e:
            logger.warning(f"Error máscara zonas: {e}")

    # Calcular área real de cada píxel en m²
    # Aproximación: 1 grado lat ≈ 111320m, 1 grado lon ≈ 111320 * cos(lat) m
    bbox_floats_calc = list(map(float, bbox.split(",")))
    min_lon_c, min_lat_c, max_lon_c, max_lat_c = bbox_floats_calc
    lat_center = (min_lat_c + max_lat_c) / 2.0
    import math
    m_per_deg_lat = 111320.0
    m_per_deg_lon = 111320.0 * math.cos(math.radians(lat_center))
    bbox_width_m  = (max_lon_c - min_lon_c) * m_per_deg_lon
    bbox_height_m = (max_lat_c - min_lat_c) * m_per_deg_lat
    bbox_area_m2  = bbox_width_m * bbox_height_m
    img_h, img_w  = ndvi_arr.shape
    total_pixels  = img_w * img_h
    pixel_area_m2_real = bbox_area_m2 / total_pixels
    logger.info(f"bbox área: {bbox_area_m2:.1f}m², píxel real: {pixel_area_m2_real:.2f}m²")

    # Interpolar píxeles sin datos (-999) usando vecinos válidos - numpy puro
    sin_datos = ndvi_arr <= -999
    if sin_datos.any() and (~sin_datos).any():
        ndvi_temp = ndvi_arr.copy()
        # Hacer hasta 3 pasadas para rellenar huecos progresivamente
        for _ in range(3):
            sin_datos_iter = ndvi_temp <= -999
            if not sin_datos_iter.any():
                break
            # Desplazamientos en las 8 direcciones
            for dy, dx in [(-1,0),(1,0),(0,-1),(0,1),(-1,-1),(-1,1),(1,-1),(1,1)]:
                shifted = np.roll(np.roll(ndvi_temp, dy, axis=0), dx, axis=1)
                # Rellenar píxeles inválidos con vecino válido
                mask_fill = sin_datos_iter & (shifted > -999)
                ndvi_temp[mask_fill] = shifted[mask_fill]
                sin_datos_iter = ndvi_temp <= -999
                if not sin_datos_iter.any():
                    break
        ndvi_arr = ndvi_temp
        sin_datos = ndvi_arr <= -999
        logger.info(f"Píxeles interpolados. Restantes sin datos: {sin_datos.sum()}")

    # Crear imagen coloreada por zonas
    h, w = ndvi_arr.shape
    rgb_arr = np.zeros((h, w, 3), dtype=np.uint8)
    rgb_arr[:] = [30, 30, 30]  # Fondo oscuro (fuera de parcela)

    zonas_stats = []

    for zona in ZONAS_NDVI:
        in_zone = (ndvi_arr >= zona["min"]) & (ndvi_arr <= zona["max"]) & (ndvi_arr > -999)
        if mask_arr is not None:
            in_zone = in_zone & mask_arr

        pixel_count = int(in_zone.sum())
        # Área real de cada píxel según bbox y tamaño de imagen
        superficie_m2 = pixel_count * pixel_area_m2_real
        superficie_ha = superficie_m2 / M2_PER_HA

        rgb_arr[in_zone] = zona["color"]

        zonas_stats.append({
            "zona": zona["zona"],
            "label": zona["label"],
            "ndvi_min": zona["min"],
            "ndvi_max": zona["max"],
            "color_hex": "#{:02x}{:02x}{:02x}".format(*zona["color"]),
            "pixeles": pixel_count,
            "superficie_m2": float(round(superficie_m2, 1)),
            "superficie_ha": float(round(superficie_ha, 4)),
        })

    # Fuera de parcela: negro
    if mask_arr is not None:
        rgb_arr[~mask_arr] = [15, 15, 15]
    # Píxeles sin datos dentro de parcela: gris oscuro
    sin_datos = (ndvi_arr <= -999)
    if mask_arr is not None:
        sin_datos = sin_datos & mask_arr
    rgb_arr[sin_datos] = [40, 40, 40]

    img_result = Image.fromarray(rgb_arr, mode="RGB")
    buf = io.BytesIO()
    img_result.save(buf, format="JPEG", quality=92)
    img_bytes = buf.getvalue()

    cache_img.write_bytes(img_bytes)
    cache_data.write_text(json.dumps(zonas_stats))

    return JSONResponse(content={
        "imagen_url": f"/ndvi/zonas/imagen?ck={ck}",
        "zonas": zonas_stats,
    })


@app.get("/ndvi/zonas/imagen")
async def ndvi_zonas_imagen(ck: str = Query(...)):
    """Sirve la imagen de zonas NDVI cacheada."""
    cache_img = CACHE_DIR / f"{ck}_zonas.jpg"
    if not cache_img.exists():
        raise HTTPException(status_code=404, detail="Imagen no encontrada, recalcula primero")
    return StreamingResponse(io.BytesIO(cache_img.read_bytes()), media_type="image/jpeg")


@app.post("/ndvi/produccion")
async def calcular_produccion(
    datos: dict,
):
    """
    Calcula kg esperados por zona y total de parcela.
    Body: { "zonas": [...], "kg_por_ha": { "1": 5000, "2": 4500, ... } }
    """
    zonas = datos.get("zonas", [])
    kg_por_ha = datos.get("kg_por_ha", {})

    resultado = []
    total_kg = 0.0
    total_ha = 0.0

    for zona in zonas:
        zona_id = str(zona["zona"])
        sup_ha = zona.get("superficie_ha", 0)
        kg_ha = float(kg_por_ha.get(zona_id, 0))
        kg_zona = sup_ha * kg_ha

        resultado.append({
            "zona": zona["zona"],
            "label": zona["label"],
            "ndvi_min": zona["ndvi_min"],
            "ndvi_max": zona["ndvi_max"],
            "color_hex": zona["color_hex"],
            "superficie_ha": round(sup_ha, 4),
            "kg_por_ha": kg_ha,
            "kg_estimados": round(kg_zona, 1),
        })
        total_kg += kg_zona
        total_ha += sup_ha

    return {
        "zonas": resultado,
        "total_ha": round(total_ha, 4),
        "total_kg": round(total_kg, 1),
        "total_toneladas": round(total_kg / 1000, 3),
    }



# ── ANÁLISIS DE FLUJO DE AGUA ─────────────────────────────────────────────

IGN_WCS_MDT = "https://servicios.idee.es/wcs-inspire/mdt"


async def descargar_mdt(bbox: str, resolucion: int = 5) -> Optional[np.ndarray]:
    """
    Descarga el MDT del IGN via WCS y lo devuelve como array numpy.
    resolucion: 5 (5m/px) o 2 (2m/px)
    """
    min_lon, min_lat, max_lon, max_lat = map(float, bbox.split(","))

    # Calcular tamaño de imagen según resolución y bbox
    import math
    m_per_deg_lat = 111320.0
    m_per_deg_lon = 111320.0 * math.cos(math.radians((min_lat + max_lat) / 2))
    width_m  = (max_lon - min_lon) * m_per_deg_lon
    height_m = (max_lat - min_lat) * m_per_deg_lat
    width_px  = max(64, min(1024, int(width_m / resolucion)))
    height_px = max(64, min(1024, int(height_m / resolucion)))

    params = {
        "SERVICE": "WCS",
        "VERSION": "1.0.0",
        "REQUEST": "GetCoverage",
        "COVERAGE": "Elevacion4258_5",
        "CRS": "EPSG:4258",
        "BBOX": f"{min_lon},{min_lat},{max_lon},{max_lat}",
        "WIDTH": str(width_px),
        "HEIGHT": str(height_px),
        "FORMAT": "image/tiff",
        "RESX": str((max_lon - min_lon) / width_px),
        "RESY": str((max_lat - min_lat) / height_px),
    }

    try:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            resp = await client.get(IGN_WCS_MDT, params=params)
            logger.info(f"WCS MDT status: {resp.status_code}, size: {len(resp.content)}")
            if resp.status_code != 200:
                return None

            # Decodificar GeoTIFF con PIL (sin rasterio)
            img = Image.open(io.BytesIO(resp.content))
            arr = np.array(img, dtype=np.float32)

            # Si es RGB, convertir a escala de grises
            if arr.ndim == 3:
                arr = arr.mean(axis=2)

            # Reemplazar nodata (-9999, 0 extremos) con nan
            arr[arr < -100] = np.nan
            arr[arr > 9000] = np.nan

            logger.info(f"MDT descargado: {arr.shape}, min={np.nanmin(arr):.1f}m, max={np.nanmax(arr):.1f}m")
            return arr

    except Exception as e:
        logger.error(f"Error descargando MDT: {e}")
        return None


def calcular_pendiente(dem: np.ndarray, res_m: float = 5.0) -> np.ndarray:
    """Calcula pendiente en grados usando diferencias finitas."""
    # Gradientes en x e y
    dz_dx = np.gradient(dem, res_m, axis=1)
    dz_dy = np.gradient(dem, res_m, axis=0)
    pendiente_rad = np.arctan(np.sqrt(dz_dx**2 + dz_dy**2))
    return np.degrees(pendiente_rad)


def calcular_direccion_flujo_d8(dem: np.ndarray) -> np.ndarray:
    """
    Algoritmo D8: para cada píxel, determina hacia cuál de los 8 vecinos
    fluye el agua (el de mayor pendiente descendente).
    Valores: 1=E, 2=SE, 4=S, 8=SW, 16=W, 32=NW, 64=N, 128=NE
    """
    h, w = dem.shape
    direction = np.zeros((h, w), dtype=np.uint8)

    # Rellenar NaN con interpolación simple
    dem_filled = dem.copy()
    nan_mask = np.isnan(dem_filled)
    if nan_mask.any():
        from numpy.lib.stride_tricks import sliding_window_view
        for _ in range(5):
            nan_mask = np.isnan(dem_filled)
            if not nan_mask.any():
                break
            for dy, dx in [(-1,0),(1,0),(0,-1),(0,1)]:
                shifted = np.roll(np.roll(dem_filled, dy, axis=0), dx, axis=1)
                fill = nan_mask & ~np.isnan(shifted)
                dem_filled[fill] = shifted[fill]

    # D8: 8 direcciones con sus pesos (diagonales * sqrt(2))
    dirs = [
        (0,  1,  1.0,  1),   # E
        (1,  1,  1.414, 2),  # SE
        (1,  0,  1.0,  4),   # S
        (1, -1,  1.414, 8),  # SW
        (0, -1,  1.0,  16),  # W
        (-1,-1,  1.414, 32), # NW
        (-1, 0,  1.0,  64),  # N
        (-1, 1,  1.414, 128),# NE
    ]

    for i in range(1, h-1):
        for j in range(1, w-1):
            z0 = dem_filled[i, j]
            max_slope = -np.inf
            best_dir = 1
            for dy, dx, dist, d_val in dirs:
                z1 = dem_filled[i+dy, j+dx]
                slope = (z0 - z1) / dist
                if slope > max_slope:
                    max_slope = slope
                    best_dir = d_val
            direction[i, j] = best_dir if max_slope > 0 else 0

    return direction


def calcular_acumulacion_flujo(direction: np.ndarray) -> np.ndarray:
    """
    Calcula la acumulación de flujo: cuántos píxeles aguas arriba
    drenan hacia cada punto. Zonas de alta acumulación = cauces.
    """
    h, w = direction.shape
    accumulation = np.ones((h, w), dtype=np.int32)

    # Mapeo de dirección → desplazamiento (dy, dx) del receptor
    dir_to_delta = {
        1: (0, 1), 2: (1, 1), 4: (1, 0), 8: (1, -1),
        16: (0,-1), 32: (-1,-1), 64: (-1, 0), 128: (-1, 1)
    }

    # Ordenar píxeles por elevación descendente (procesar cimas primero)
    dem_flat = direction  # usamos direction como proxy
    order = np.argsort(direction, axis=None)[::-1]

    for idx in order:
        i, j = divmod(int(idx), w)
        d = direction[i, j]
        if d == 0:
            continue
        if d in dir_to_delta:
            di, dj = dir_to_delta[d]
            ni, nj = i + di, j + dj
            if 0 <= ni < h and 0 <= nj < w:
                accumulation[ni, nj] += accumulation[i, j]

    return accumulation


def detectar_valles(dem: np.ndarray, mask_arr) -> np.ndarray:
    dem_f = np.nan_to_num(dem, nan=float(np.nanmean(dem)))
    curv_x = np.gradient(np.gradient(dem_f, axis=1), axis=1)
    curv_y = np.gradient(np.gradient(dem_f, axis=0), axis=0)
    valle = (curv_x > 0) | (curv_y > 0)
    dz_dx = np.gradient(dem_f, axis=1)
    dz_dy = np.gradient(dem_f, axis=0)
    pendiente = np.sqrt(dz_dx**2 + dz_dy**2)
    pend_norm = pendiente / (pendiente.max() + 1e-10)
    valle = valle & (pend_norm < 0.85)
    if mask_arr is not None:
        valle = valle & mask_arr
    return valle


def trazar_lineas_flujo(direction: np.ndarray, acumulacion: np.ndarray,
                        dem: np.ndarray, mask_arr, max_lineas: int = 120) -> list:
    h, w = direction.shape
    dir_to_delta = {
        1: (0, 1), 2: (1, 1), 4: (1, 0), 8: (1, -1),
        16: (0,-1), 32: (-1,-1), 64: (-1, 0), 128: (-1, 1)
    }
    acc_norm = np.log1p(acumulacion.astype(np.float32))
    acc_max = acc_norm.max()
    if acc_max > 0:
        acc_norm = acc_norm / acc_max

    valles = detectar_valles(dem, mask_arr)
    umbral_acc = np.percentile(acc_norm, 70)
    puntos_inicio = np.argwhere(valles & (acc_norm >= umbral_acc))

    if len(puntos_inicio) == 0:
        umbral_acc = np.percentile(acc_norm, 80)
        puntos_inicio = np.argwhere(acc_norm >= umbral_acc)

    if len(puntos_inicio) > max_lineas:
        idx = np.round(np.linspace(0, len(puntos_inicio)-1, max_lineas)).astype(int)
        puntos_inicio = puntos_inicio[idx]

    lineas = []
    visitados_global = set()

    for pi, pj in puntos_inicio:
        ci, cj = int(pi), int(pj)
        linea = [(ci, cj)]
        visitados_linea = set()
        pasos = 0
        while pasos < 800:
            key = (ci, cj)
            if key in visitados_global or key in visitados_linea:
                break
            visitados_linea.add(key)
            d = direction[ci, cj]
            if d == 0 or d not in dir_to_delta:
                break
            di, dj = dir_to_delta[d]
            ni, nj = ci + di, cj + dj
            if not (0 <= ni < h and 0 <= nj < w):
                break
            if mask_arr is not None and not mask_arr[ni, nj]:
                break
            linea.append((ni, nj))
            ci, cj = ni, nj
            pasos += 1
        if len(linea) > 8:
            for p in linea[len(linea)//2:]:
                visitados_global.add(p)
            lineas.append(linea)
    return lineas


def dibujar_linea_suave(rgb: np.ndarray, linea: list, color: tuple, grosor: int = 1):
    h, w = rgb.shape[:2]
    for k in range(len(linea) - 1):
        i0, j0 = linea[k]
        i1, j1 = linea[k+1]
        di = abs(i1 - i0); dj = abs(j1 - j0)
        si = 1 if i0 < i1 else -1
        sj = 1 if j0 < j1 else -1
        err = di - dj
        ci, cj = i0, j0
        while True:
            for gi in range(-grosor, grosor+1):
                for gj in range(-grosor, grosor+1):
                    ni, nj = ci+gi, cj+gj
                    if 0 <= ni < h and 0 <= nj < w:
                        rgb[ni, nj] = color
            if ci == i1 and cj == j1:
                break
            e2 = 2 * err
            if e2 > -dj:
                err -= dj; ci += si
            if e2 < di:
                err += di; cj += sj


def generar_imagen_flujo(
    dem: np.ndarray,
    acumulacion: np.ndarray,
    pendiente: np.ndarray,
    direction: np.ndarray,
    mask_arr = None,
) -> bytes:
    """
    Genera imagen JPEG con hillshade realista y líneas de flujo de agua (D8 + detección de valles).
    """
    h, w = dem.shape
    rgb = np.zeros((h, w, 3), dtype=np.uint8)

    # Hillshade realista
    dem_f = dem.copy()
    dem_f = np.nan_to_num(dem_f, nan=float(np.nanmean(dem_f)))
    dz_dy, dz_dx = np.gradient(dem_f)
    azimuth = np.radians(315)
    altitud  = np.radians(45)
    normal_x = -dz_dx; normal_y = -dz_dy; normal_z = np.ones_like(dem_f)
    mag = np.sqrt(normal_x**2 + normal_y**2 + normal_z**2)
    normal_x /= mag; normal_y /= mag; normal_z /= mag
    luz_x = np.cos(altitud) * np.cos(azimuth)
    luz_y = np.cos(altitud) * np.sin(azimuth)
    luz_z = np.sin(altitud)
    hillshade = np.clip(normal_x * luz_x + normal_y * luz_y + normal_z * luz_z, 0, 1)

    dmin, dmax = np.nanmin(dem_f), np.nanmax(dem_f)
    elev_norm = (dem_f - dmin) / (dmax - dmin + 1e-10)
    r_terr = (180 + elev_norm * 60).astype(np.uint8)
    g_terr = (160 + elev_norm * 40).astype(np.uint8)
    b_terr = (120 + elev_norm * 20).astype(np.uint8)
    rgb[:,:,0] = np.clip(r_terr * hillshade * 1.1, 0, 255).astype(np.uint8)
    rgb[:,:,1] = np.clip(g_terr * hillshade * 1.1, 0, 255).astype(np.uint8)
    rgb[:,:,2] = np.clip(b_terr * hillshade * 1.1, 0, 255).astype(np.uint8)

    if mask_arr is not None:
        rgb[~mask_arr] = [30, 30, 35]

    # Líneas de flujo
    acc_norm = np.log1p(acumulacion.astype(np.float32))
    acc_max = acc_norm.max()
    if acc_max > 0:
        acc_norm = acc_norm / acc_max

    lineas = trazar_lineas_flujo(direction, acumulacion, dem_f, mask_arr)

    def acc_media(linea):
        vals = [acc_norm[i, j] for i, j in linea]
        return sum(vals) / len(vals)

    lineas_sorted = sorted(lineas, key=acc_media)
    umbral_cauce = np.percentile(acc_norm, 90)
    umbral_afluente = np.percentile(acc_norm, 75)

    for linea in lineas_sorted:
        am = acc_media(linea)
        if am >= umbral_cauce:
            color = (20, 80, 200); grosor = 2
        elif am >= umbral_afluente:
            color = (60, 140, 220); grosor = 1
        else:
            color = (120, 180, 240); grosor = 1
        dibujar_linea_suave(rgb, linea, color, grosor)

    img = Image.fromarray(rgb, mode='RGB')
    buf = io.BytesIO()
    img.save(buf, format='JPEG', quality=90)
    return buf.getvalue()

@app.get("/flujo/analizar")
async def analizar_flujo(
    bbox: str = Query(...),
    geojson: Optional[str] = Query(None),
):
    """
    Descarga MDT del IGN, calcula dirección y acumulación de flujo (D8)
    y devuelve imagen con cauces y zonas de acumulación de agua.
    """
    ck = cache_key("flujo16", bbox=bbox, mask=bool(geojson))
    cache_img  = CACHE_DIR / f"{ck}_flujo.jpg"
    cache_data = CACHE_DIR / f"{ck}_flujo.json"

    if cache_img.exists() and cache_data.exists():
        return JSONResponse(content={
            "imagen_url": f"/flujo/imagen?ck={ck}",
            "stats": json.loads(cache_data.read_text()),
        })

    # Descargar MDT
    dem = await descargar_mdt(bbox, resolucion=5)
    if dem is None:
        raise HTTPException(status_code=502, detail="No se pudo descargar el MDT del IGN. Inténtalo de nuevo.")

    # Aplicar máscara de parcela
    mask_arr = None
    if geojson:
        try:
            geojson_data = json.loads(geojson)
            bbox_floats = list(map(float, bbox.split(",")))
            mask = geojson_to_mask(geojson_data, dem.shape[1], dem.shape[0], bbox_floats)
            mask_arr = np.array(mask) > 128
        except Exception as e:
            logger.warning(f"Error máscara flujo: {e}")

    # Calcular pendiente
    pendiente = calcular_pendiente(dem, res_m=5.0)

    # Calcular dirección y acumulación de flujo D8
    logger.info("Calculando dirección de flujo D8...")
    direction = calcular_direccion_flujo_d8(dem)
    logger.info("Calculando acumulación de flujo...")
    acumulacion = calcular_acumulacion_flujo(direction)

    # Estadísticas dentro de parcela
    if mask_arr is not None:
        dem_parcela   = dem[mask_arr]
        pend_parcela  = pendiente[mask_arr]
        acum_parcela  = acumulacion[mask_arr]
    else:
        dem_parcela   = dem.flatten()
        pend_parcela  = pendiente.flatten()
        acum_parcela  = acumulacion.flatten()

    dem_valido = dem_parcela[~np.isnan(dem_parcela)]
    stats = {
        "altitud_min":   round(float(np.nanmin(dem_valido)), 1) if len(dem_valido) > 0 else 0,
        "altitud_max":   round(float(np.nanmax(dem_valido)), 1) if len(dem_valido) > 0 else 0,
        "altitud_media": round(float(np.nanmean(dem_valido)), 1) if len(dem_valido) > 0 else 0,
        "pendiente_media": round(float(np.nanmean(pend_parcela[~np.isnan(pend_parcela)])), 1),
        "pendiente_max":   round(float(np.nanmax(pend_parcela[~np.isnan(pend_parcela)])), 1),
        "pixeles_cauce": int((acum_parcela >= np.percentile(acumulacion, 95)).sum()),
    }

    # Generar imagen
    img_bytes = generar_imagen_flujo(dem, acumulacion, pendiente, direction, mask_arr)

    cache_img.write_bytes(img_bytes)
    cache_data.write_text(json.dumps(stats))

    return JSONResponse(content={
        "imagen_url": f"/flujo/imagen?ck={ck}",
        "stats": stats,
    })


@app.get("/flujo/imagen")
async def flujo_imagen(ck: str = Query(...)):
    """Sirve la imagen de flujo cacheada."""
    cache_img = CACHE_DIR / f"{ck}_flujo.jpg"
    if not cache_img.exists():
        raise HTTPException(status_code=404, detail="Imagen no encontrada, recalcula primero")
    return StreamingResponse(io.BytesIO(cache_img.read_bytes()), media_type="image/jpeg")



# ── MRL DATABASE ───────────────────────────────────────────────────────────────
_mrl_cache: dict = {}
_mrl_cargado: bool = False
_mrl_fecha_carga: str = ""

# Ruta a los Excel de fitosanitarios incluidos en el repo
MRL_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "lmr_materias_activas_espana.xlsx")
MRL_FITO_PATH  = os.path.join(os.path.dirname(__file__), "fitosanitarios_espana.xlsx")


async def cargar_mrl_database() -> bool:
    global _mrl_cache, _mrl_cargado, _mrl_fecha_carga
    try:
        import openpyxl
        nueva = {}

        if os.path.exists(MRL_EXCEL_PATH):
            wb = openpyxl.load_workbook(MRL_EXCEL_PATH, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) >= 3:
                header = [str(c).lower().strip() if c else "" for c in rows[2]]
                def fc(kws):
                    for kw in kws:
                        for i, h in enumerate(header):
                            if kw in h: return i
                    return -1
                idx_c = fc(["cultivo"]); idx_s = fc(["materia activa","sustancia"]); idx_m = fc(["lmr","mg/kg","limite"])
                if idx_s >= 0 and idx_c >= 0 and idx_m >= 0:
                    for row in rows[3:]:
                        if not row or not row[idx_s] or not row[idx_c]: continue
                        s = str(row[idx_s]).strip(); c = str(row[idx_c]).strip()
                        m = str(row[idx_m]).strip() if row[idx_m] else "Sin LMR"
                        if s and c:
                            nueva[(s.lower(), c.lower())] = {"sustancia_oficial": s, "cultivo_oficial": c, "mrl": m, "unidad": "mg/kg"}
            wb.close()
            logger.info(f"MRL LMR Excel: {len(nueva)} entradas")

        if os.path.exists(MRL_FITO_PATH):
            wb2 = openpyxl.load_workbook(MRL_FITO_PATH, read_only=True, data_only=True)
            ws2 = wb2.active
            rows2 = list(ws2.iter_rows(values_only=True))
            if len(rows2) >= 3:
                h2 = [str(c).lower().strip() if c else "" for c in rows2[2]]
                def fc2(kws):
                    for kw in kws:
                        for i, h in enumerate(h2):
                            if kw in h: return i
                    return -1
                idx_s2 = fc2(["materia activa"]); idx_c2 = fc2(["cultivo"])
                if idx_s2 >= 0 and idx_c2 >= 0:
                    for row in rows2[3:]:
                        if not row or not row[idx_s2] or not row[idx_c2]: continue
                        s = str(row[idx_s2]).strip(); c = str(row[idx_c2]).strip()
                        key = (s.lower(), c.lower())
                        if key not in nueva:
                            nueva[key] = {"sustancia_oficial": s, "cultivo_oficial": c, "mrl": "Consultar etiqueta", "unidad": "mg/kg"}
            wb2.close()
            logger.info(f"MRL total con fitosanitarios: {len(nueva)} entradas")

        if len(nueva) > 0:
            _mrl_cache = nueva; _mrl_cargado = True
            _mrl_fecha_carga = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
            logger.info(f"BD MRL lista: {len(_mrl_cache)} entradas")
            return True
    except Exception as e:
        logger.error(f"Error cargando MRL desde Excel: {e}")

    logger.warning("No se pudo cargar BD MRL")
    return False
@app.on_event("startup")
async def startup_event():
    await cargar_mrl_database()


@app.get("/mrl/consultar")
async def consultar_mrl(
    sustancia: str = Query(..., description="Materia activa"),
    cultivo: str = Query(..., description="Cultivo de la parcela"),
):
    if not _mrl_cargado:
        await cargar_mrl_database()
    resultado = buscar_mrl(sustancia, cultivo)
    return {
        "sustancia_consultada": sustancia,
        "cultivo_consultado": cultivo,
        "base_datos": "EU Pesticides Database (Reg. EC 396/2005)",
        "fecha_carga": _mrl_fecha_carga,
        "total_entradas_bd": len(_mrl_cache),
        **resultado,
    }


@app.get("/mrl/estado")
async def mrl_estado():
    return {"cargado": _mrl_cargado, "entradas": len(_mrl_cache), "fecha_carga": _mrl_fecha_carga}


@app.post("/mrl/recargar")
async def mrl_recargar():
    global _mrl_cargado
    _mrl_cargado = False
    ok = await cargar_mrl_database()
    return {"ok": ok, "entradas": len(_mrl_cache)}



# -- SIAR (Sistema de Informacion Agroclimatica para el Regadio) ---------------
SIAR_BASE  = "https://servicio.mapa.gob.es/siarapi"
SIAR_USER  = os.getenv("SIAR_USER", "")   # NIF/usuario REGEUS
SIAR_PASS  = os.getenv("SIAR_PASS", "")   # Contrasena REGEUS
_siar_token: str = ""
_siar_token_ts: float = 0.0
_siar_estaciones: list = []
_siar_estaciones_cargadas: bool = False


async def _siar_obtener_token() -> str:
    """Obtiene token SIAR cifrando usuario y contrasena con la API de SIAR."""
    global _siar_token, _siar_token_ts
    import time
    # Reusar token si tiene menos de 8 horas
    if _siar_token and (time.time() - _siar_token_ts) < 28800:
        return _siar_token
    if not SIAR_USER or not SIAR_PASS:
        logger.warning("SIAR: SIAR_USER o SIAR_PASS no configurados")
        return ""
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            r1 = await client.get(f"{SIAR_BASE}/API/V1/Autenticacion/cifrarCadena?cadena={SIAR_USER}")
            logger.info(f"SIAR cifrar usuario: {r1.status_code} | {r1.text[:80]}")
            usuario_cifrado = r1.text.strip().strip('"') or ""

            r2 = await client.get(f"{SIAR_BASE}/API/V1/Autenticacion/cifrarCadena?cadena={SIAR_PASS}")
            logger.info(f"SIAR cifrar pass: {r2.status_code} | {r2.text[:80]}")
            pass_cifrado = r2.text.strip().strip('"') or ""

            if not usuario_cifrado or not pass_cifrado:
                logger.error("SIAR: cifrado vacio")
                return ""

            r3 = await client.get(f"{SIAR_BASE}/API/V1/Autenticacion/obtenerToken?Usuario={usuario_cifrado}&Password={pass_cifrado}")
            logger.info(f"SIAR token: {r3.status_code} | {r3.text[:200]}")
            token = r3.text.strip().strip('"') or ""
            if len(token) > 10:
                _siar_token = token
                _siar_token_ts = time.time()
                logger.info(f"SIAR: token OK (...{token[-8:]})")
                return _siar_token
            else:
                logger.error(f"SIAR: token inesperado: {repr(token)}")
    except Exception as ex:
        logger.error(f"Error token SIAR: {ex}")
    return ""



async def _cargar_estaciones_siar() -> list:
    global _siar_estaciones, _siar_estaciones_cargadas
    if _siar_estaciones_cargadas and _siar_estaciones:
        return _siar_estaciones
    token = await _siar_obtener_token()
    if not token:
        return []
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(f"{SIAR_BASE}/API/V1/Info/ESTACIONES", params={"token": token})
            r.raise_for_status()
            estaciones_raw = r.json().get("datos", [])
        logger.info(f"SIAR: {len(estaciones_raw)} estaciones en catalogo raw")
        resultado = []
        for e in estaciones_raw:
            try:
                # Fecha_Baja = null -> activa; Fecha_Baja = fecha -> dada de baja
                if e.get("Fecha_Baja") is not None:
                    continue
                lat = _parse_siar_coord(str(e.get("Latitud", "")))
                lon = _parse_siar_coord(str(e.get("Longitud", "")))
                if abs(lat) < 0.1 and abs(lon) < 0.1:
                    continue
                if not (27 <= lat <= 44 and -19 <= lon <= 5):
                    continue
                resultado.append({
                    "id": e.get("Codigo", ""),
                    "nombre": e.get("Estacion", ""),
                    "termino": e.get("Termino", ""),
                    "lat": lat, "lon": lon,
                })
            except Exception:
                continue
        _siar_estaciones = resultado
        _siar_estaciones_cargadas = True
        logger.info(f"SIAR: {len(_siar_estaciones)} estaciones cargadas")
        return _siar_estaciones
    except Exception as ex:
        logger.error(f"Error estaciones SIAR: {ex}")
        return []


async def _datos_siar_mensual(estacion_id: str, fecha_ini: str, fecha_fin: str) -> list:
    """Obtiene datos MENSUALES del SIAR para acumulados (menos celdas que diario)."""
    token = await _siar_obtener_token()
    if not token:
        return []
    try:
        params = {
            "token": token,
            "Id": estacion_id,
            "FechaInicial": fecha_ini,
            "FechaFinal": fecha_fin,
            "DatosCalculados": "true",
        }
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(f"{SIAR_BASE}/API/V1/Datos/Mensuales/ESTACION", params=params)
            if r.status_code in (400, 403):
                logger.warning(f"SIAR mensual {r.status_code} {estacion_id}: {r.text[:100]}")
                return []
            r.raise_for_status()
            return r.json().get("datos", []) or []
    except Exception as ex:
        logger.error(f"Error datos SIAR mensual: {ex}")
        return []


async def _datos_siar_periodo(estacion_id: str, fecha_ini: str, fecha_fin: str) -> list:
    token = await _siar_obtener_token()
    if not token:
        return []
    try:
        params = {
            "token": token,
            "Id": estacion_id,
            "FechaInicial": fecha_ini,
            "FechaFinal": fecha_fin,
            "DatosCalculados": "true",
        }
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.get(f"{SIAR_BASE}/API/V1/Datos/Diarios/ESTACION", params=params)
            if r.status_code in (400, 403):
                logger.warning(f"SIAR {r.status_code} {estacion_id}: {r.text[:100]}")
                return []
            r.raise_for_status()
            return r.json().get("datos", []) or []
    except Exception as ex:
        logger.error(f"Error datos SIAR: {ex}")
        return []

async def _clima_desde_siar(lat: float, lon: float) -> dict:
    """
    Una sola peticion al SIAR para todo el año agricola.
    Extrae datos del ultimo dia y acumula el resto.
    Evita el rate limit haciendo solo 1 consulta en vez de 2.
    """
    from datetime import date, timedelta
    estaciones = await _cargar_estaciones_siar()
    if not estaciones:
        return {}
    estacion = min(estaciones, key=lambda e: _haversine(lat, lon, e["lat"], e["lon"]))

    hoy = date.today()
    if hoy.month >= 10:
        inicio_agricola = date(hoy.year, 10, 1)
    else:
        inicio_agricola = date(hoy.year - 1, 10, 1)

    import asyncio

    def num(v, default=None):
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return default

    # Peticion 1: datos del ultimo dia (temperatura + precip + ETo)
    ayer = hoy - timedelta(days=1)
    datos_dia = await _datos_siar_periodo(
        estacion["id"],
        ayer.strftime("%Y-%m-%d"),
        hoy.strftime("%Y-%m-%d")
    )

    if not datos_dia:
        return {}

    obs = datos_dia[-1]
    temp_max   = num(obs.get("TempMax"))
    temp_min   = num(obs.get("TempMin"))
    temp_med   = num(obs.get("TempMedia"))
    precip_dia = num(obs.get("Precipitacion"), 0.0)
    eto_dia    = num(obs.get("EtPMon"))

    if eto_dia is None and temp_max is not None and temp_min is not None:
        lat_rad = _math.radians(lat)
        dia = ayer.timetuple().tm_yday
        tmedia = temp_med if temp_med is not None else (temp_max + temp_min) / 2
        eto_dia = _eto_hargreaves(temp_max, temp_min, tmedia, lat_rad, dia)

    # Esperar 1s para no superar el rate limit
    await asyncio.sleep(1.2)

    # Peticion 2: datos MENSUALES para acumulados del año agricola (max ~10 meses = 10 filas)
    datos_mes = await _datos_siar_mensual(
        estacion["id"],
        inicio_agricola.strftime("%Y-%m-%d"),
        hoy.strftime("%Y-%m-%d")
    )

    precip_acum = sum(num(d.get("Precipitacion"), 0.0) or 0 for d in datos_mes)
    eto_acum    = sum(num(d.get("EtPMon"), 0.0) or 0 for d in datos_mes)

    # Si mensual no devuelve datos, intentar con campos alternativos
    if not datos_mes:
        precip_acum = None
        eto_acum    = None

    return {
        "fuente": "SIAR",
        "estacion": estacion["nombre"],
        "termino": estacion.get("termino", ""),
        "temperatura_actual": round(temp_med, 1) if temp_med is not None else (round((temp_max+temp_min)/2, 1) if temp_max and temp_min else None),
        "temperatura_maxima": round(temp_max, 1) if temp_max is not None else None,
        "temperatura_minima": round(temp_min, 1) if temp_min is not None else None,
        "precipitacion_dia": round(precip_dia, 1) if precip_dia is not None else 0.0,
        "eto_dia": round(eto_dia, 2) if eto_dia is not None else None,
        "precipitacion_anyo_agricola": round(precip_acum, 1),
        "eto_anyo_agricola": round(eto_acum, 1),
    }


# ── CLIMATOLOGÍA AEMET ────────────────────────────────────────────────────────

import math as _math

def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distancia en km entre dos puntos geográficos."""
    R = 6371.0
    dlat = _math.radians(lat2 - lat1)
    dlon = _math.radians(lon2 - lon1)
    a = _math.sin(dlat/2)**2 + _math.cos(_math.radians(lat1)) * _math.cos(_math.radians(lat2)) * _math.sin(dlon/2)**2
    return R * 2 * _math.asin(_math.sqrt(a))

def _centroide(geojson: dict) -> tuple:
    """Calcula el centroide de un GeoJSON."""
    try:
        geom = geojson["features"][0]["geometry"]
        coords = []
        if geom["type"] == "Polygon":
            coords = geom["coordinates"][0]
        elif geom["type"] == "MultiPolygon":
            for poly in geom["coordinates"]:
                coords.extend(poly[0])
        if not coords:
            return None, None
        lons = [c[0] for c in coords]
        lats = [c[1] for c in coords]
        return sum(lats)/len(lats), sum(lons)/len(lons)
    except Exception:
        return None, None

def _eto_hargreaves(tmax: float, tmin: float, tmedia: float, lat_rad: float, dia_anyo: int) -> float:
    """
    ETo por método de Hargreaves-Samani (mm/día).
    Solo necesita Tmax, Tmin, Tmedia y latitud.
    """
    try:
        dr = 1 + 0.033 * _math.cos(2 * _math.pi * dia_anyo / 365)
        decl = 0.409 * _math.sin(2 * _math.pi * dia_anyo / 365 - 1.39)
        ws = _math.acos(-_math.tan(lat_rad) * _math.tan(decl))
        Ra = (24*60/_math.pi) * 0.0820 * dr * (ws * _math.sin(lat_rad) * _math.sin(decl) + _math.cos(lat_rad) * _math.cos(decl) * _math.sin(ws))
        eto = 0.0023 * Ra * (tmedia + 17.8) * (tmax - tmin)**0.5
        return round(max(0, eto), 2)
    except Exception:
        return None

# Cache de estaciones AEMET en memoria
_aemet_estaciones: list = []
_aemet_estaciones_cargadas: bool = False

# Estaciones AEMET de respaldo (principales de España) para cuando la API falla
_ESTACIONES_FALLBACK = [
    {"id":"5402","lat":40.4165,"lon":-3.7026},   # Madrid Retiro
    {"id":"B013X","lat":41.3888,"lon":2.1590},   # Barcelona
    {"id":"6155A","lat":37.3773,"lon":-5.9810},  # Sevilla
    {"id":"7031","lat":37.1667,"lon":-3.5833},   # Granada
    {"id":"7178I","lat":36.6667,"lon":-4.5500},  # Malaga
    {"id":"7031X","lat":37.8500,"lon":-4.7667},  # Cordoba
    {"id":"6001","lat":37.3827,"lon":-5.9736},   # Sevilla aerop.
    {"id":"3129","lat":39.4667,"lon":-0.3667},   # Valencia
    {"id":"2462","lat":41.6500,"lon":-0.8833},   # Zaragoza
    {"id":"1387","lat":43.3500,"lon":-8.4167},   # A Coruña
    {"id":"1082","lat":43.4667,"lon":-1.7833},   # San Sebastian
    {"id":"1024E","lat":43.5500,"lon":-5.8833},  # Oviedo
    {"id":"9001","lat":41.6500,"lon":-4.7667},   # Valladolid
    {"id":"4121","lat":38.9833,"lon":-3.9167},   # Ciudad Real
    {"id":"8175","lat":38.9833,"lon":-1.1333},   # Albacete
    {"id":"8096","lat":37.6833,"lon":-0.8000},   # Murcia
    {"id":"9434","lat":39.8833,"lon":-4.0500},   # Toledo
    {"id":"2539","lat":40.9667,"lon":-5.5000},   # Salamanca
    {"id":"2914","lat":40.4833,"lon":-3.5500},   # Madrid Barajas
    {"id":"6293X","lat":36.7500,"lon":-6.1333},  # Jerez
    {"id":"5783","lat":39.8833,"lon":-4.0500},   # Badajoz
    {"id":"1208H","lat":43.3000,"lon":-1.9833},  # Pamplona
    {"id":"9898","lat":40.9500,"lon":-5.5000},   # Salamanca 2
    {"id":"B954","lat":39.5500,"lon":2.6167},    # Palma Mallorca
    {"id":"C449C","lat":28.4500,"lon":-13.8667}, # Fuerteventura
    {"id":"C249I","lat":28.9667,"lon":-13.5500}, # Lanzarote
    {"id":"C447A","lat":27.9167,"lon":-15.3833}, # Gran Canaria
    {"id":"C029O","lat":28.0500,"lon":-16.5667}, # Tenerife Sur
]

async def _cargar_estaciones_aemet() -> list:
    """Descarga y cachea el catálogo de estaciones AEMET. Usa fallback si falla."""
    global _aemet_estaciones, _aemet_estaciones_cargadas
    if _aemet_estaciones_cargadas and _aemet_estaciones:
        return _aemet_estaciones
    if AEMET_API_KEY:
        try:
            headers = {"api_key": AEMET_API_KEY, "Accept": "application/json"}
            logger.info(f"AEMET: consultando catálogo con key ...{AEMET_API_KEY[-10:]}")
            async with httpx.AsyncClient(timeout=60) as client:
                r = await client.get(f"{AEMET_BASE}/valores/climatologicos/inventarioestaciones/todasestaciones", headers=headers)
                logger.info(f"AEMET catálogo status: {r.status_code}")
                resp_json = r.json()
                logger.info(f"AEMET catálogo respuesta: {resp_json}")
                r.raise_for_status()
                data_url = resp_json.get("datos")
                if data_url:
                    logger.info(f"AEMET descargando datos desde: {data_url}")
                    r2 = await client.get(data_url, timeout=60)
                    r2.raise_for_status()
                    # AEMET devuelve latin-1, no UTF-8
                    estaciones = json.loads(r2.content.decode('latin-1'))
                    logger.info(f"AEMET: {len(estaciones)} estaciones en catálogo raw")
                    resultado = []
                    # Log first station to see coordinate format
                    if estaciones:
                        logger.info(f"AEMET ejemplo estacion: {estaciones[0]}")
                    errores = 0
                    for e in estaciones:
                        try:
                            lat_s = str(e.get("latitud", "")).strip()
                            lon_s = str(e.get("longitud", "")).strip()
                            if not lat_s or not lon_s:
                                continue
                            lat = _parse_dms(lat_s)
                            lon = _parse_dms(lon_s)
                            # Filtrar coordenadas claramente inválidas
                            if abs(lat) < 0.1 and abs(lon) < 0.1:
                                errores += 1
                                continue
                            # Filtrar fuera de España+Canarias
                            if not (27 <= lat <= 44 and -19 <= lon <= 5):
                                continue
                            resultado.append({
                                "id": e["indicativo"],
                                "lat": lat,
                                "lon": lon,
                                "nombre": e.get("nombre", "").strip(),
                                "municipio": e.get("municipio", "").strip(),
                            })
                        except Exception:
                            errores += 1
                            continue
                    logger.info(f"AEMET: {len(resultado)} estaciones válidas, {errores} descartadas")
                    if resultado:
                        _aemet_estaciones = resultado
                        _aemet_estaciones_cargadas = True
                        logger.info(f"AEMET: {len(_aemet_estaciones)} estaciones cargadas")
                        return _aemet_estaciones
        except Exception as ex:
            logger.warning(f"Error cargando estaciones AEMET (usando fallback): {ex}")

    # Usar estaciones de respaldo
    logger.info(f"AEMET: usando {len(_ESTACIONES_FALLBACK)} estaciones de respaldo")
    _aemet_estaciones = _ESTACIONES_FALLBACK
    _aemet_estaciones_cargadas = True
    return _aemet_estaciones

def _parse_siar_coord(s: str) -> float:
    """
    Parsea coordenadas en formato SIAR: DDMMSSMMM + direccion
    Ejemplo: '015512000W' -> 1°55'12" W -> -1.9200
             '391520000N' -> 39°15'20" N -> 39.2556
    Los últimos 3 dígitos son milisegundos (se ignoran).
    """
    s = str(s).strip()
    if not s:
        return 0.0
    neg = s[-1] in ('S', 'W', 'O')
    s = s[:-1]
    # Quitar los 3 últimos dígitos (milisegundos)
    core = s[:-3] if len(s) >= 9 else s
    try:
        if len(core) == 6:
            dd, mm, ss = int(core[0:2]), int(core[2:4]), int(core[4:6])
        elif len(core) == 7:
            dd, mm, ss = int(core[0:3]), int(core[3:5]), int(core[5:7])
        else:
            return float(core) * (-1 if neg else 1)
        val = dd + mm/60.0 + ss/3600.0
        return -val if neg else val
    except Exception:
        return 0.0


def _parse_dms(s: str) -> float:
    """
    Convierte coordenadas AEMET a decimal.
    Formatos posibles:
    - '412429N'  → 41° 24' 29" N  → 41.408056
    - '0024500W' → 2° 45' 00" W   → -2.75
    - '41.4081'  → decimal directo
    - '41,4081'  → decimal con coma
    """
    s = str(s).strip()
    if not s:
        return 0.0
    # Detectar hemisferio
    neg = False
    if s[-1] in ('S', 'W', 'O'):
        neg = True
    if s[-1] in ('N', 'S', 'E', 'W', 'O'):
        s = s[:-1]
    # Reemplazar coma decimal
    s = s.replace(",", ".")
    try:
        # Formato decimal
        if "." in s:
            val = float(s)
            return -val if neg else val
        # Formato DDMMSS o DDDMMSS (sin punto)
        s_digits = s.lstrip("0") or "0"
        num = int(s)
        # Determinar si es DDMMSS (6 dígitos) o DDDMMSS (7 dígitos)
        if len(s) >= 7:
            # Longitud: DDDMMSS
            dd = int(s[:-4])
            mm = int(s[-4:-2])
            ss = int(s[-2:])
        elif len(s) >= 5:
            # Latitud: DDMMSS
            dd = int(s[:-4])
            mm = int(s[-4:-2])
            ss = int(s[-2:])
        else:
            dd = int(s)
            mm = 0
            ss = 0
        val = dd + mm/60.0 + ss/3600.0
        return -val if neg else val
    except Exception as ex:
        return 0.0

async def _estacion_mas_cercana(lat: float, lon: float) -> str:
    """Devuelve el indicativo AEMET de la estación más cercana."""
    estaciones = await _cargar_estaciones_aemet()
    if not estaciones:
        return None
    mejor = min(estaciones, key=lambda e: _haversine(lat, lon, e["lat"], e["lon"]))
    return mejor["id"]

async def _obs_estacion(indicativo: str) -> dict:
    """Obtiene observaciones actuales de una estación AEMET."""
    try:
        headers = {"api_key": AEMET_API_KEY, "Accept": "application/json"}
        async with httpx.AsyncClient(timeout=20) as client:
            r = await client.get(f"{AEMET_BASE}/observacion/convencional/datos/estacion/{indicativo}", headers=headers)
            r.raise_for_status()
            data_url = r.json().get("datos")
            if not data_url:
                return {}
            r2 = await client.get(data_url)
            r2.raise_for_status()
            obs_list = r2.json()
            if not obs_list:
                return {}
            obs = obs_list[-1]
            return obs
    except Exception as ex:
        logger.error(f"Error obs AEMET {indicativo}: {ex}")
        return {}

async def _nombre_estacion(indicativo: str) -> str:
    """Obtiene el nombre legible de la estación AEMET desde el catálogo cargado."""
    # Buscar en el catálogo cargado (que ya tiene nombres si se cargó bien)
    for e in _aemet_estaciones:
        if e.get("id") == indicativo:
            nombre = e.get("nombre", "")
            municipio = e.get("municipio", "")
            if nombre:
                return f"{nombre} ({municipio})" if municipio else nombre
    return indicativo

async def _eto_anyo_agricola(indicativo: str, lat: float) -> float:
    """ETo acumulada desde 1 oct del año agrícola actual (Hargreaves diario)."""
    try:
        from datetime import date
        hoy = date.today()
        if hoy.month >= 10:
            inicio = date(hoy.year, 10, 1)
        else:
            inicio = date(hoy.year - 1, 10, 1)
        fi = inicio.strftime("%Y-%m-%dT00:00:00UTC")
        ff = hoy.strftime("%Y-%m-%dT23:59:59UTC")
        headers = {"api_key": AEMET_API_KEY, "Accept": "application/json"}
        async with httpx.AsyncClient(timeout=30) as client:
            url = f"{AEMET_BASE}/valores/climatologicos/diarios/datos/fechaini/{fi}/fechafin/{ff}/estacion/{indicativo}"
            r = await client.get(url, headers=headers)
            r.raise_for_status()
            data_url = r.json().get("datos")
            if not data_url:
                return None
            r2 = await client.get(data_url)
            r2.raise_for_status()
            datos = json.loads(r2.content.decode('latin-1'))
        lat_rad = _math.radians(lat)
        total_eto = 0.0
        for d in datos:
            try:
                tmax = float(str(d.get("tmax","")).replace(",","."))
                tmin = float(str(d.get("tmin","")).replace(",","."))
                tmedia = (tmax + tmin) / 2
                # Día del año
                fecha_str = d.get("fecha", "")
                if fecha_str:
                    from datetime import datetime
                    dia = datetime.strptime(fecha_str, "%Y-%m-%dT%H:%M:%S").timetuple().tm_yday
                else:
                    dia = 180
                eto_dia = _eto_hargreaves(tmax, tmin, tmedia, lat_rad, dia)
                if eto_dia:
                    total_eto += eto_dia
            except Exception:
                continue
        return round(total_eto, 1)
    except Exception as ex:
        logger.error(f"Error ETo año agrícola: {ex}")
        return None


async def _precip_anyo_agricola(indicativo: str, lat: float) -> float:
    """Precipitación acumulada desde 1 oct del año agrícola actual."""
    try:
        from datetime import date, timedelta
        hoy = date.today()
        # Año agrícola: empieza 1 octubre
        if hoy.month >= 10:
            inicio = date(hoy.year, 10, 1)
        else:
            inicio = date(hoy.year - 1, 10, 1)
        fi = inicio.strftime("%Y-%m-%dT00:00:00UTC")
        ff = hoy.strftime("%Y-%m-%dT23:59:59UTC")
        headers = {"api_key": AEMET_API_KEY, "Accept": "application/json"}
        async with httpx.AsyncClient(timeout=30) as client:
            url = f"{AEMET_BASE}/valores/climatologicos/diarios/datos/fechaini/{fi}/fechafin/{ff}/estacion/{indicativo}"
            r = await client.get(url, headers=headers)
            r.raise_for_status()
            data_url = r.json().get("datos")
            if not data_url:
                return None
            r2 = await client.get(data_url)
            r2.raise_for_status()
            datos = json.loads(r2.content.decode('latin-1'))
        total = 0.0
        for d in datos:
            p = str(d.get("prec","0")).replace(",",".").replace("Ip","0")
            try:
                total += float(p)
            except Exception:
                pass
        return round(total, 1)
    except Exception as ex:
        logger.error(f"Error precip año agrícola: {ex}")
        return None


@app.get("/clima/parcela")
async def clima_parcela(
    lat: float = Query(..., description="Latitud del centroide"),
    lon: float = Query(..., description="Longitud del centroide"),
):
    """
    Devuelve datos climatológicos actuales.
    Prioridad 1: SIAR (datos agronómicos, ETo Penman-Monteith, acumulados año agrícola)
    Prioridad 2: AEMET (fallback si SIAR no disponible)
    Los acumulados (precip + ETo año agrícola) siempre vienen del SIAR si está disponible.
    """
    from datetime import date, timedelta

    # --- INTENTAR SIAR PRIMERO ---
    siar_datos = {}
    siar_disponible = bool(SIAR_USER and SIAR_PASS)
    if siar_disponible:
        try:
            siar_datos = await _clima_desde_siar(lat, lon)
        except Exception as ex:
            logger.warning(f"SIAR falló, usando AEMET: {ex}")

    if siar_datos and siar_datos.get("temperatura_actual") is not None:
        # SIAR OK - devolver directamente con acumulados incluidos
        return {
            "fuente": "SIAR",
            "estacion": siar_datos.get("estacion", ""),
            "temperatura_actual": siar_datos.get("temperatura_actual"),
            "temperatura_maxima": siar_datos.get("temperatura_maxima"),
            "temperatura_minima": siar_datos.get("temperatura_minima"),
            "precipitacion_dia": siar_datos.get("precipitacion_dia"),
            "precipitacion_anyo_agricola": siar_datos.get("precipitacion_anyo_agricola"),
            "eto_dia": siar_datos.get("eto_dia"),
            "eto_anyo_agricola": siar_datos.get("eto_anyo_agricola"),
            "unidades": {"temperatura": "°C", "precipitacion": "mm", "eto": "mm/dia"}
        }

    # --- FALLBACK: AEMET ---
    logger.info("Usando AEMET como fuente de datos climatológicos")
    if not AEMET_API_KEY:
        raise HTTPException(status_code=503, detail="Sin datos climatológicos disponibles")

    indicativo = await _estacion_mas_cercana(lat, lon)
    if not indicativo:
        indicativo = "5402"

    obs = await _obs_estacion(indicativo)

    def _num(val, default=None):
        if val is None:
            return default
        try:
            return float(str(val).replace(",", "."))
        except Exception:
            return default

    temp_actual = _num(obs.get("ta"))
    temp_max    = _num(obs.get("tamax") or obs.get("ta"))
    temp_min    = _num(obs.get("tamin") or obs.get("ta"))
    precip_dia  = _num(obs.get("prec"), 0.0)
    if temp_max is None: temp_max = temp_actual
    if temp_min is None: temp_min = temp_actual

    eto = None
    if temp_max is not None and temp_min is not None:
        lat_rad = _math.radians(lat)
        dia = date.today().timetuple().tm_yday
        tmedia = (temp_max + temp_min) / 2
        eto = _eto_hargreaves(temp_max, temp_min, tmedia, lat_rad, dia)

    nombre_estacion = await _nombre_estacion(indicativo)

    # Acumulados: intentar desde SIAR aunque los datos actuales vengan de AEMET
    precip_anyo = None
    eto_anyo = None
    if siar_disponible and siar_datos:
        precip_anyo = siar_datos.get("precipitacion_anyo_agricola")
        eto_anyo    = siar_datos.get("eto_anyo_agricola")
    else:
        precip_anyo = await _precip_anyo_agricola(indicativo, lat)
        eto_anyo    = await _eto_anyo_agricola(indicativo, lat)

    return {
        "fuente": "AEMET",
        "estacion": nombre_estacion,
        "temperatura_actual": round(temp_actual, 1) if temp_actual is not None else None,
        "temperatura_maxima": round(temp_max, 1) if temp_max is not None else None,
        "temperatura_minima": round(temp_min, 1) if temp_min is not None else None,
        "precipitacion_dia": round(precip_dia, 1) if precip_dia is not None else None,
        "precipitacion_anyo_agricola": precip_anyo,
        "eto_dia": eto,
        "eto_anyo_agricola": eto_anyo,
        "unidades": {"temperatura": "°C", "precipitacion": "mm", "eto": "mm/dia"}
    }


@app.post("/aemet/recargar-estaciones")
async def recargar_estaciones_aemet():
    """Fuerza la recarga del catálogo de estaciones AEMET."""
    global _aemet_estaciones_cargadas, _aemet_estaciones
    _aemet_estaciones_cargadas = False
    _aemet_estaciones = []
    estaciones = await _cargar_estaciones_aemet()
    return {"ok": True, "total_estaciones": len(estaciones)}


# -- REGISTRO PRODUCTOS FITOSANITARIOS (MAPA) ----------------------------------
_fito_productos: list = []    # Vigentes
_fito_eliminados: list = []   # Historial eliminados
_fito_cargado: bool = False

FITO_JSON_PATH       = os.path.join(os.path.dirname(__file__), "ProductosAutorizados.json")
FITO_ELIMINADOS_PATH = os.path.join(os.path.dirname(__file__), "productos_eliminados.json")
FITO_PDF_DIRS = [
    os.path.join(os.path.dirname(__file__), "pdfs-1"),
    os.path.join(os.path.dirname(__file__), "pdfs-2"),
    os.path.join(os.path.dirname(__file__), "pdfs-3"),
]

def _buscar_pdf(id_producto: int):
    for carpeta in FITO_PDF_DIRS:
        ruta = os.path.join(carpeta, f"{id_producto}.pdf")
        if os.path.exists(ruta):
            return ruta
    return None

def _tiene_pdf(id_producto: int) -> bool:
    return _buscar_pdf(id_producto) is not None


def _cargar_fito():
    global _fito_productos, _fito_eliminados, _fito_cargado
    if _fito_cargado:
        return
    try:
        # Cargar productos vigentes
        if os.path.exists(FITO_JSON_PATH):
            with open(FITO_JSON_PATH, "r", encoding="utf-8-sig") as f:
                data = json.load(f)
            _fito_productos = [
                {
                    "id": p["IdProducto"],
                    "nombre": p.get("Nombre", "").strip(),
                    "num_registro": p.get("NumRegistro", ""),
                    "titular": p.get("Titular", ""),
                    "formulado": p.get("Formulado", ""),
                    "estado": "Vigente",
                    "fecha_caducidad": p.get("StrFechaCaducidad", ""),
                    "observaciones": (p.get("Condicionamiento") or p.get("Observaciones") or "")[:300],
                    "eliminado": False,
                    "fecha_eliminacion": None,
                }
                for p in data if p.get("Nombre")
            ]
            logger.info(f"Fitosanitarios vigentes: {len(_fito_productos)}")

        # Cargar historial eliminados
        if os.path.exists(FITO_ELIMINADOS_PATH):
            with open(FITO_ELIMINADOS_PATH, "r", encoding="utf-8") as f:
                data_el = json.load(f)
            _fito_eliminados = [
                {
                    "id": p["IdProducto"],
                    "nombre": p.get("Nombre", "").strip(),
                    "num_registro": p.get("NumRegistro", ""),
                    "titular": p.get("Titular", ""),
                    "formulado": p.get("Formulado", ""),
                    "estado": "Eliminado",
                    "fecha_caducidad": "",
                    "observaciones": p.get("Observaciones", "")[:300],
                    "eliminado": True,
                    "fecha_eliminacion": p.get("FechaEliminacion", ""),
                }
                for p in data_el if p.get("Nombre")
            ]
            logger.info(f"Fitosanitarios eliminados: {len(_fito_eliminados)}")

        _fito_cargado = True
    except Exception as ex:
        logger.error(f"Error cargando fitosanitarios: {ex}")



@app.on_event("startup")
async def startup_fito():
    _cargar_fito()


@app.get("/fito/buscar")
async def fito_buscar(q: str = Query(..., min_length=2)):
    """Busca productos fitosanitarios (vigentes + eliminados) por nombre o nº registro."""
    _cargar_fito()
    q_low = q.lower().strip()
    todos = _fito_productos + _fito_eliminados
    resultados = [
        {**p, "tiene_pdf": _tiene_pdf(p["id"])}
        for p in todos
        if q_low in p["nombre"].lower() or q_low in p["num_registro"].lower()
    ][:20]
    return {"total": len(resultados), "productos": resultados}


@app.get("/fito/producto/{id_producto}")
async def fito_detalle(id_producto: int):
    """Devuelve datos del producto. Incluye URL del PDF si existe."""
    _cargar_fito()
    todos = _fito_productos + _fito_eliminados
    producto = next((p for p in todos if p["id"] == id_producto), None)
    if not producto:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    tiene_pdf = _tiene_pdf(id_producto)
    return {
        **producto,
        "tiene_pdf": tiene_pdf,
        "pdf_url": f"/fito/pdf/{id_producto}" if tiene_pdf else None,
        "ficha_web_url": f"https://servicio.mapa.gob.es/regfiweb/BuscadorProductos/Index?nombre={producto['nombre']}",
    }


@app.get("/fito/pdf/{id_producto}")
async def fito_pdf(id_producto: int):
    """Sirve el PDF de la ficha tecnica del producto."""
    from fastapi.responses import FileResponse
    pdf_path = _buscar_pdf(id_producto)
    if not pdf_path:
        raise HTTPException(status_code=404, detail="PDF no disponible")
    return FileResponse(pdf_path, media_type="application/pdf",
                        filename=f"ficha_{id_producto}.pdf")


@app.get("/cache/info")
async def cache_info():
    files = list(CACHE_DIR.glob("*"))
    total_mb = sum(f.stat().st_size for f in files if f.is_file()) / 1e6
    return {"archivos": len(files), "total_mb": round(total_mb, 2)}


@app.delete("/cache/limpiar")
async def limpiar_cache(dias: int = Query(7)):
    cutoff = time.time() - dias * 86400
    eliminados = sum(1 for f in CACHE_DIR.glob("*") if f.is_file() and f.stat().st_mtime < cutoff and not f.unlink())
    return {"eliminados": eliminados}
